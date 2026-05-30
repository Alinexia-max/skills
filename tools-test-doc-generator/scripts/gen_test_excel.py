# -*- coding: utf-8 -*-
"""
研发自测用例Excel生成器
用法：
  1. 将测试用例数据填入下方 cases 列表
  2. 设置 OUTPUT_FILE 为目标输出路径
  3. 执行 python gen_test_excel.py

依赖：pip install openpyxl
"""
import os
import sys
import traceback

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ======================== 配置区 ========================
# 输出文件路径（由调用方设置）
# 命名规则：{日期}_test_{单号}_开发自测报告_通过.xlsx，日期格式yyyyMMdd
# 示例：20260526_test_ERR260416115_开发自测报告_通过.xlsx
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "研发自测文档.xlsx")

# 测试用例数据 - 每行格式：
# [用例编号, 用例名称, 前提条件, 测试步骤, 预期结果, 备注, 用例等级, 用例类型, 需求编号, 功能模块, 功能子模块, 结果]
# 由调用方填入实际数据
cases = [
    # 示例数据，实际使用时替换
    [1, "示例用例", "无", "1. 执行操作", "预期结果", "", "P0", "功能测试", "待补充", "示例模块", "示例子模块", ""],
]

# ======================== 生成逻辑 ========================

def generate(output_file, test_cases):
    """生成测试用例Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "自测用例"

    # 样式定义
    hfont = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cfont = Font(name="Microsoft YaHei", size=10)
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_a = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    # 表头
    headers = ["用例编号", "用例名称", "前提条件", "测试步骤", "预期结果", "备注", "用例等级", "用例类型", "需求编号", "功能模块", "功能子模块", "结果"]
    col_widths = [8, 22, 28, 40, 40, 12, 8, 10, 10, 14, 14, 8]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = center_a
        cell.border = thin
        # 设置列宽
        col_letter = chr(64 + ci) if ci <= 26 else chr(64 + ci // 26) + chr(65 + ci % 26)
        ws.column_dimensions[col_letter].width = w

    # 填充数据
    for r, case in enumerate(test_cases, 2):
        for ci, val in enumerate(case, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = cfont
            cell.alignment = wrap if ci >= 3 else center_a
            cell.border = thin

    # 冻结首行（不启用自动筛选）
    ws.freeze_panes = "A2"

    # 确保输出目录存在
    out_dir = os.path.dirname(os.path.abspath(output_file))
    if not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_file)
    print("SUCCESS: " + os.path.abspath(output_file))
    return output_file


if __name__ == "__main__":
    try:
        result = generate(OUTPUT_FILE, cases)
    except Exception as e:
        print("FAIL: " + str(e))
        traceback.print_exc()
        sys.exit(1)
